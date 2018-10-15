

from openpyxl import Workbook
#import process_data
import cx_Oracle as oracle 
import datetime
import sqls

def write_excel(user,passwd):
	#create a workbook
	wb = Workbook()
	#create a sheet
	sheet1 = wb.create_sheet(index=0,title='outresource')
	results_all = get_test_datas(user,passwd)
	#results_all = get_pro_datas()
	print('add title...')
	sheet1.append(sqls.W_TITLE)
	print('add lines...')
	for i in results_all:
		sheet1.append(i)

    #sheet1.add_table()
	print('save excel...')
	name_str = 'w_resouece_'+str(datetime.datetime.now().strftime('%Y%m%d%H%I%S'))+'.xlsx'
	wb.save(name_str)
	return name_str

def get_test_datas(user,passwd):
	#tns
	tns = oracle.makedsn('172.18.2.152',1531,'terp1')
	#connection
	db = oracle.connect(user,passwd,tns,encoding='utf-8',nencoding='utf-8')
	#cursor
	cursor = db.cursor()
	print('cursor...')

	#execute sql
	print('cursor execute...')
	cursor.execute(sqls.W_RESOURCE,sqls.W_PARAMS)
	#fetch results
	result_all = cursor.fetchall()
	#close cursor
	cursor.close()
	#close connection
	db.close()
	print('db.close()...')
	return result_all

class WriteExcel(Workbook):
	def __init__(self,user,passwd):
		print('__init__')
		super().__init__() #初始化基类 在此派生类的函数中就可以直接调用基类函数了
		self.user = user
		self.passwd = passwd
		
	#设置通用型变量
	def set_infos(self,data_kind,data_name,data_title):
		self.data_kind = data_kind
		self.data_name = data_name
		self.data_title = data_title
	
	def write_excel(self):
		#create a sheet
		sheet1 = self.create_sheet(index=0,title=str(self.data_name))
		results_all = get_test_datas(self.user,self.passwd)
		#results_all = get_pro_datas()
		print('add title...')
		sheet1.append(self.data_title)
		print('add lines...')
		for i in results_all:
			sheet1.append(i)
		#sheet1.add_table()
		print('save excel...')
		name_str = str(self.data_name)+str(datetime.datetime.now().strftime('%Y%m%d%H%I%S'))+'.xlsx'
		self.save(name_str)
		return name_str

